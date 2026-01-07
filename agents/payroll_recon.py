import os
import glob
import pandas as pd
import pdfplumber
import re
import openpyxl
import io
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# --- Google Drive Imports ---
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# =================CONFIGURATION=================
# Get the project root directory (parent of Agent folder)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

# Local Folders (relative to project root)
INPUT_FOLDER = os.path.join(PROJECT_ROOT, 'Payroll Recon')
OUTPUT_FOLDER = os.path.join(PROJECT_ROOT, 'Payroll Recon Code and Output')

# File Names
TEMPLATE_FILENAME = 'Payroll Reconciliation ZPI - 10.31.25.xlsx'
OUTPUT_FILENAME = 'Reconciled_Output.xlsx'
NEW_SHEET_NAME = 'Reconciliation -Workday'

# Database Log (relative to project root)
DB_PATH = os.path.join(PROJECT_ROOT, "config_mapping.xlsx")
DB_SHEET_NAME = "Payroll_Agent_Database"

# Google Drive Configuration (relative to project root)
SERVICE_ACCOUNT_FILE = os.path.join(PROJECT_ROOT, "service_account.json")
DRIVE_FOLDER_ID = '10l-6EkqzO3J3aaWRGZGvPDmhxvNwxw6G'  # Folder ID from your prompt
SCOPES = ['https://www.googleapis.com/auth/drive']
# ===============================================

# --- 1. GOOGLE DRIVE FUNCTIONS ---

def get_drive_service():
    """Authenticates with Google Drive using Service Account."""
    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        print(f"⚠️ Warning: '{SERVICE_ACCOUNT_FILE}' not found. Skipping Drive download.")
        return None
    
    try:
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES
        )
        return build('drive', 'v3', credentials=creds)
    except Exception as e:
        print(f"❌ Drive Auth Error: {e}")
        return None

def download_files_from_drive(service, folder_id, local_dir):
    """Downloads ALL files from the specific Drive folder to local_dir."""
    if not service: return

    if not os.path.exists(local_dir):
        os.makedirs(local_dir)

    print(f"🔍 Checking Google Drive Folder ID: {folder_id}...")
    
    try:
        # List files in the specific folder
        query = f"'{folder_id}' in parents and trashed = false"
        results = service.files().list(
            q=query, fields="files(id, name, mimeType)",
            supportsAllDrives=True, includeItemsFromAllDrives=True
        ).execute()
        
        files = results.get('files', [])

        if not files:
            print("   -> No files found in Drive folder.")
            return

        print(f"   -> Found {len(files)} files. Downloading...")

        for file in files:
            file_id = file['id']
            file_name = file['name']
            
            # Skip folders, only download files
            if file['mimeType'] == 'application/vnd.google-apps.folder':
                continue

            local_path = os.path.join(local_dir, file_name)
            
            # Download
            request = service.files().get_media(fileId=file_id)
            with io.FileIO(local_path, 'wb') as fh:
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()
            
            print(f"      ⬇️ Downloaded: {file_name}")

    except Exception as e:
        print(f"❌ Drive Download Error: {e}")

# --- 2. DATA PROCESSING HELPERS ---

def clean_currency(value_str):
    if isinstance(value_str, (int, float)): return float(value_str)
    try:
        val_str = str(value_str).strip()
        is_negative = '-' in val_str or ('(' in val_str and ')' in val_str)
        clean = re.sub(r'[^\d.]', '', val_str)
        val = float(clean) if clean else 0.0
        return -val if is_negative else val
    except: return 0.0

def find_header_row_dynamically(file_path, sheet_name, required_keywords):
    try:
        df_temp = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=30)
        for idx, row in df_temp.iterrows():
            row_str = " ".join(row.astype(str).values).lower()
            if sheet_name == 'Netsuite Reports':
                match = all(k.lower() in row_str for k in required_keywords)
            else:
                match = any(k.lower() in row_str for k in required_keywords)
            if match: return idx, pd.read_excel(file_path, sheet_name=sheet_name, header=idx)
        return None, None
    except: return None, None

def extract_net_pay_from_pdf(pdf_path):
    net_pay = 0.0
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[-1].extract_text()
            match = re.search(r'(?:Net Pay|Direct Deposit).*?([\d,]+\.\d{2})', text, re.IGNORECASE)
            if match:
                net_pay = clean_currency(match.group(1))
            else:
                for line in text.split('\n'):
                    if "Total" in line and ("Net" in line or "Deposit" in line):
                        nums = re.findall(r'[\d,]+\.\d{2}', line)
                        if nums: net_pay = clean_currency(nums[-1])
    except: pass
    return abs(net_pay)

# --- 3. DATABASE LOGGING ---

def update_database_log(output_file_path):
    print(f"\nStep 5: Updating Database at {DB_PATH}...")
    db_dir = os.path.dirname(DB_PATH)
    if not os.path.exists(db_dir): os.makedirs(db_dir, exist_ok=True)

    if os.path.exists(DB_PATH):
        try: wb = openpyxl.load_workbook(DB_PATH)
        except: 
            print("  -> Error loading DB. Is it open?")
            return
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]

    if DB_SHEET_NAME in wb.sheetnames: ws = wb[DB_SHEET_NAME]
    else:
        ws = wb.create_sheet(DB_SHEET_NAME)
        headers = ["Date", "Processed Timestamp", "Output File Name", "Output File Link", "Approval Status"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center")

    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=datetime.now().strftime("%Y-%m-%d"))
    ws.cell(row=r, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(row=r, column=3, value=os.path.basename(output_file_path))
    
    link_cell = ws.cell(row=r, column=4, value="Click to Open")
    link_cell.hyperlink = os.path.abspath(output_file_path)
    link_cell.font = Font(color="0563C1", underline="single")

    status_cell = ws.cell(row=r, column=5, value="Pending")
    dv = DataValidation(type="list", formula1='"Pending,Approved,Rejected"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(status_cell)

    try: wb.save(DB_PATH); print("  -> Database Updated.")
    except Exception as e: print(f"  -> Save Error: {e}")

# --- 4. EXCEL BUILDER ---

def build_tables(wb, recon_map, pdf_details):
    if NEW_SHEET_NAME in wb.sheetnames:
        ws = wb[NEW_SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
    else: ws = wb.create_sheet(NEW_SHEET_NAME)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws['A1'] = "Source File Name"; ws['B1'] = "Workday Net Pay"
    for c in [ws['A1'], ws['B1']]: 
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, align_center, thin_border

    r = 2
    for item in pdf_details:
        ws.cell(r, 1, item['Source File']).border = thin_border
        c2 = ws.cell(r, 2, item['Workday Net Pay'])
        c2.number_format = '"$"#,##0.00'; c2.border = thin_border
        r += 1
    ws.column_dimensions['A'].width = 45; ws.column_dimensions['B'].width = 20

    ws['E1'] = "Reconciliation Item"; ws['F1'] = "Amount"
    for c in [ws['E1'], ws['F1']]:
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, align_center, thin_border

    data = [
        ("Total per Workday", recon_map['Total per Workday'], False),
        ("Total per Report (NetSuite 21110)", recon_map['Total per Report'], False),
        ("Variance (Net Pay)", "=F2-F3", True),
        ("", "", False),
        ("Taxes per Workday (NetSuite 21120)", recon_map['Taxes per Workday'], False),
        ("Tax Collection (OSV)", recon_map['Tax Collection'], False),
        ("Variance (Taxes)", "=F6-F7", True)
    ]
    
    r = 2
    for label, val, hl in data:
        c_l, c_v = ws.cell(r, 5, label), ws.cell(r, 6, val)
        c_l.border, c_v.border = thin_border, thin_border
        if hl:
            c_l.fill = c_v.fill = PatternFill("solid", fgColor="FFF2CC")
            c_l.font = c_v.font = Font(bold=True)
        if val != "": c_v.number_format = '"$"#,##0.00'
        r += 1
    ws.column_dimensions['E'].width = 40; ws.column_dimensions['F'].width = 20
    return wb

# --- 5. MAIN EXECUTION ---

def run_payroll_agent():
    print("🚀 Starting Payroll Reconciliation Agent v16.0 (Google Drive Enabled)")
    
    if not os.path.exists(OUTPUT_FOLDER): os.makedirs(OUTPUT_FOLDER)
    
    # --- STEP 1: DOWNLOAD FROM DRIVE ---
    print("\nStep 1: Connecting to Google Drive...")
    drive_service = get_drive_service()
    if drive_service:
        download_files_from_drive(drive_service, DRIVE_FOLDER_ID, INPUT_FOLDER)
    else:
        print("   -> Proceeding with existing local files only.")

    # --- STEP 2: GATHER PDF DATA ---
    print("\nStep 2: Identifying Files...")
    pdf_files = glob.glob(os.path.join(INPUT_FOLDER, '*.pdf'))
    pdf_details = []
    total_workday = 0.0
    
    for pdf in pdf_files:
        amt = extract_net_pay_from_pdf(pdf)
        pdf_details.append({'Source File': os.path.basename(pdf), 'Workday Net Pay': amt})
        total_workday += amt

    # --- STEP 3: PROCESS EXCEL ---
    print("\nStep 3: Processing Excel Data...")
    template_path = os.path.join(INPUT_FOLDER, TEMPLATE_FILENAME)
    total_report = 0.0; taxes_workday = 0.0; tax_collection = 0.0
    
    if os.path.exists(template_path):
        # NetSuite
        idx, df_ns = find_header_row_dynamically(template_path, 'Netsuite Reports', ['Account', 'Net'])
        if df_ns is not None:
            df_ns.columns = df_ns.columns.astype(str).str.strip()
            col_acc = next((c for c in df_ns.columns if 'Account' in c), None)
            col_net = next((c for c in df_ns.columns if 'Net' in c or 'Amount' in c), None)
            if col_acc and col_net:
                df_ns[col_net] = df_ns[col_net].apply(clean_currency)
                df_ns[col_acc] = df_ns[col_acc].astype(str)
                total_report = abs(df_ns.loc[df_ns[col_acc].str.contains('21110', na=False), col_net].sum())
                taxes_workday = abs(df_ns.loc[df_ns[col_acc].str.contains('21120', na=False), col_net].sum())

        # OSV
        idx, df_osv = find_header_row_dynamically(template_path, 'OSV', ['Type', 'Total', 'Amount'])
        if df_osv is not None:
            df_osv.columns = df_osv.columns.astype(str).str.strip()
            col_type = next((c for c in df_osv.columns if 'type' in c.lower()), None)
            col_total = next((c for c in df_osv.columns if 'total' in c.lower() or 'amount' in c.lower()), None)
            if col_type and col_total:
                df_osv[col_total] = df_osv[col_total].apply(clean_currency)
                mask = df_osv[col_type].astype(str).str.contains('Tax Collection', case=False, na=False)
                tax_collection = abs(df_osv.loc[mask, col_total].sum())
    else:
        print(f"⚠️ Warning: Excel Template '{TEMPLATE_FILENAME}' not found in {INPUT_FOLDER}")

    # --- STEP 4: GENERATE OUTPUT ---
    print("\nStep 4: Generating Output File...")
    recon_map = {
        'Total per Workday': total_workday, 'Total per Report': total_report,
        'Taxes per Workday': taxes_workday, 'Tax Collection': tax_collection
    }
    
    out_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)
    try:
        if os.path.exists(template_path): wb = openpyxl.load_workbook(template_path)
        else: wb = openpyxl.Workbook()
        wb = build_tables(wb, recon_map, pdf_details)
        for sheet_name in ['Workday', 'Copy of Workday']:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
        wb.create_sheet('Adjustment')
        ws_adj = wb['Adjustment']
        headers = ['Debit', 'Credit', 'Date', 'Memo', 'Account', 'Department', 'Class', 'Location', 'External ID', 'Subsidiary']
        ws_adj.append(headers)
        wb.save(out_path)
        print(f"✅ Success: Reconciled file saved at {out_path}")
        update_database_log(out_path)
    except Exception as e:
        print(f"❌ Error: {e}")
    return {
        "status": "completed",
        "output_file": os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)
    }

if __name__ == "__main__":
    run_payroll_agent()