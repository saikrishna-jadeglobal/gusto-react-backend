import os
import io
import pandas as pd # type: ignore
import re
import json
from datetime import datetime, timedelta
from google.oauth2 import service_account # type: ignore
from googleapiclient.discovery import build # type: ignore
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment # type: ignore
from openpyxl import load_workbook # type: ignore
from openpyxl.worksheet.datavalidation import DataValidation # type: ignore

# --- Configuration ---
SERVICE_ACCOUNT_FILE = 'service_account.json'
FOLDER_ID = '1Xqz9giE_hZYFMAcVmJSbGAItw10cPS3O' 
SCOPES = ['https://www.googleapis.com/auth/drive']

# --- File Names ---
SOURCE_FILENAME = "1 - SBC Expense FY25 10.2025 (1).xlsx"
SOURCE_TAB_RAW = "Expense (Amortization)" 
SOURCE_TAB_JE = "Expense (Journal Entries)"
TARGET_FILENAME = "1 - SBC Expense Recon FY25 10.2025 (1).xlsx"
MAPPING_FILENAME = "config_mapping.xlsx"

# Sheet Names
PIVOT_SHEET_NAME = "Pivot Data"
RECON_SHEET_NAME = "Reconciliation"
ENTRY_SHEET_NAME_US = "Entry - Gusto Inc US"
ENTRY_SHEET_NAME_CA = "Entry - Canada ($USD)"

# !!! UPDATED TO MATCH PUSH AGENT !!!
DB_SHEET_NAME = "Equity_Agent_Database"

# Directories
LOCAL_WORK_DIR = "Equity Input and Output Files"

# --- Constants ---
GROUP_BY_COLS = ['Entity', 'Dept', 'Location', 'Class']

TARGET_ENTITIES = [
    "Gusto Inc Global : Gusto Inc US",
    "Gusto Inc Global : Gusto Inc US : ZP Insurance LLC", 
    "Gusto Inc Global : Gusto Canada ULC"
]

ENTRY_ENTITIES_US = [
    "Gusto Inc Global : Gusto Inc US",
    "Gusto Inc Global : Gusto Inc US : ZP Insurance LLC"
]

ENTRY_ENTITIES_CA = [
    "Gusto Inc Global : Gusto Canada ULC"
]

# --- Helper Functions ---

def get_drive_service():
    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        raise FileNotFoundError(f"Missing {SERVICE_ACCOUNT_FILE}.")
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES
    )
    return build('drive', 'v3', credentials=creds)

def find_file_id(service, filename, folder_id):
    query = f"name = '{filename}' and '{folder_id}' in parents and trashed = false"
    results = service.files().list(
        q=query, fields="files(id, name)", 
        supportsAllDrives=True, includeItemsFromAllDrives=True
    ).execute()
    files = results.get('files', [])
    return files[0]['id'] if files else None

def list_folder_contents(service, folder_id):
    try:
        query = f"'{folder_id}' in parents and trashed = false"
        results = service.files().list(
            q=query, fields="files(id, name)", pageSize=15,
            supportsAllDrives=True, includeItemsFromAllDrives=True
        ).execute()
        files = results.get('files', [])
        print(f"\n📂 Files found in Drive ({folder_id}):")
        for f in files:
            print(f" - {f['name']}")
        print("-" * 30)
    except Exception as e:
        print(f"⚠️ Debug List Failed: {e}")

def upload_to_drive(service, local_path, folder_id, drive_filename):
    """Uploads a local file to Google Drive."""
    print(f"⬆️ Uploading '{drive_filename}' to Drive...")
    try:
        file_metadata = {'name': drive_filename, 'parents': [folder_id]}
        mimetype = 'application/json' if drive_filename.endswith('.json') else 'application/octet-stream'
        
        media = MediaFileUpload(local_path, mimetype=mimetype)
        file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        print(f"✅ Upload successful. File ID: {file.get('id')}")
    except Exception as e:
        print(f"❌ Upload failed: {e}")

def normalize_header(header):
    if not isinstance(header, str): return str(header)
    return re.sub(r'[^a-z0-9]', '', str(header).lower())

def clean_currency(series):
    s = series.astype(str).str.strip()
    s = s.str.replace('$', '', regex=False)\
         .str.replace('USD', '', regex=False)\
         .str.replace(',', '', regex=False)\
         .str.replace(' - ', '0', regex=False)
    return pd.to_numeric(s, errors='coerce').fillna(0)

# --- NetSuite Mapping & Payload Helpers ---

def load_local_mappings(mapping_path):
    if not os.path.exists(mapping_path):
        print(f"⚠️ Warning: Mapping file '{mapping_path}' not found. IDs will be 0.")
        return {}
    
    mappings = {}
    print(f"📂 Loading mappings from: {mapping_path}")
    try:
        xls = pd.ExcelFile(mapping_path)
        target_sheets = {"Account": "Account", "Dept": "Department", "Class": "Class", "Location": "Location"}

        for key, sheet_name in target_sheets.items():
            if sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                df.columns = [normalize_header(c) for c in df.columns]
                col_name = next((c for c in df.columns if 'name' in c or 'account' in c), None)
                col_id = next((c for c in df.columns if 'internalid' in c or 'id' in c), None)
                if col_name and col_id:
                    mappings[key] = pd.Series(df[col_id].values, index=df[col_name].astype(str).str.strip()).to_dict()
                else:
                    mappings[key] = {}
            else:
                mappings[key] = {}
    except Exception as e:
        print(f"❌ Error loading mappings: {e}")
        return {}
    return mappings

def get_ns_id(mapping_dict, value):
    if not value or pd.isna(value): return 0
    val_str = str(value).strip()
    if val_str in mapping_dict: return mapping_dict[val_str]
    for k, v in mapping_dict.items():
        if val_str in k or k in val_str:
            return v
    return 0

def generate_payload(df, subsidiary_id, mappings, label="Payload"):
    print(f"   ⚙️ Generating {label} from {len(df)} lines...")
    
    payload = {
        "subsidiary": subsidiary_id,
        "tran_date": (datetime.today().replace(day=1) - timedelta(days=1)).strftime('%Y-%m-%d'),
        "memo": "November 2025_SBC_US" if label == "US Payload" else "Nov 2025_SBC_CA" ,
        "lines": [],
        "approval_status": "Pending Approval"
    }
    
    if df.empty:
        print(f"   ⚠️ Warning: DataFrame for {label} is empty.")
        return payload

    for _, row in df.iterrows():
        acc_id = get_ns_id(mappings.get('Account', {}), row.get('Account'))
        dept_id = get_ns_id(mappings.get('Dept', {}), row.get('Dept'))
        class_id = get_ns_id(mappings.get('Class', {}), row.get('Class'))
        loc_id = get_ns_id(mappings.get('Location', {}), row.get('Location'))
        
        raw_debit = float(row.get('Debit', 0))
        raw_credit = float(row.get('Credit', 0))
        
        final_debit = 0
        final_credit = 0

        if raw_debit > 0.001:
            final_debit = raw_debit
            final_credit = 0
        elif raw_credit > 0.001:
            final_debit = 0
            final_credit = raw_credit
        else:
            continue 

        line_item = {
            "account": int(acc_id),
            "debit": round(final_debit, 2),
            "credit": round(final_credit, 2),
            "department": int(dept_id),
            "class": int(class_id),
            "location": int(loc_id),
            "memo": str(row.get('Description', '')),
            "entity": 0 
        }
        
        payload["lines"].append(line_item)
        
    print(f"   ✅ Generated {len(payload['lines'])} JSON lines.")
    return payload

def update_agent_database(mapping_path, us_json_path, ca_json_path, generated_excel_path):
    """
    Updates the Equity_Agent_Database sheet with the exact schema required by the Push Agent and App.
    Schema: Date | File Name | Link for file | Approval status | Payload_US | Payload_CAD | Netsuit Pushed status | Output response
    """
    print(f"🗄️ Updating '{DB_SHEET_NAME}' in local mapping file...")
    
    try:
        if not os.path.exists(mapping_path):
            print("❌ Mapping file not found.")
            return

        wb = load_workbook(mapping_path)

        if DB_SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(DB_SHEET_NAME)
            # REQUIRED HEADERS
            headers = [
                "Date", 
                "File Name", 
                "Link for file", 
                "Approval status", 
                "Payload_US", 
                "Payload_CAD", 
                "Netsuit Pushed status", 
                "Output response"
            ]
            ws.append(headers)
            
            # Formatting headers
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4472C4")
                cell.alignment = Alignment(horizontal="center")
        else:
            ws = wb[DB_SHEET_NAME]

        # Prepare Data
        current_date = datetime.now().strftime('%Y-%m-%d')
        file_name = os.path.basename(generated_excel_path)
        file_link = os.path.abspath(generated_excel_path)
        
        # Row Data
        new_row = [
            current_date,           # Date
            file_name,              # File Name
            file_link,              # Link for file
            "Pending",              # Approval status
            os.path.abspath(us_json_path), # Payload_US (Full Path)
            os.path.abspath(ca_json_path), # Payload_CAD (Full Path)
            "Pending",              # Netsuit Pushed status
            ""                      # Output response
        ]

        ws.append(new_row)
        
        # Add Hyperlink to the File Link
        last_row = ws.max_row
        link_cell = ws.cell(row=last_row, column=3)
        link_cell.hyperlink = file_link
        link_cell.font = Font(color="0563C1", underline="single")
        
        # Add Data Validation for Approval Status (Column D / 4)
        dv = DataValidation(type="list", formula1='"Pending,Approved,Rejected"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=last_row, column=4))

        wb.save(mapping_path)
        print("✅ Database updated with standardized 8-column schema.")

    except Exception as e:
        print(f"❌ Error updating database: {e}")


# --- Formatting Helpers ---
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") 
BOLD_FONT = Font(bold=True)

def apply_header_style(ws, row_idx, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

def auto_adjust_columns(ws, min_width=15):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = max(adjusted_width, min_width)

# --- Main Logic ---

def run_equity_agent( period: str, business_day: str = None, tasks: list = None) -> dict:
    print("🚀 Starting Final Automation...")

    if not os.path.exists(LOCAL_WORK_DIR): os.makedirs(LOCAL_WORK_DIR)
    local_source_path = os.path.join(LOCAL_WORK_DIR, SOURCE_FILENAME)
    local_target_path = os.path.join(LOCAL_WORK_DIR, TARGET_FILENAME)
    local_mapping_path = os.path.join(MAPPING_FILENAME)

    try:
        service = get_drive_service()
        
        # 1. Source File Logic
        file_ready = False
        if os.path.exists(local_source_path):
            print(f"📂 Found local file '{local_source_path}'. Skipping download.")
            file_ready = True
        else:
            print(f"🔍 Searching Drive Folder ({FOLDER_ID})...")
            file_id = find_file_id(service, SOURCE_FILENAME, FOLDER_ID)
            if file_id:
                print(f"⬇️ Downloading to '{local_source_path}'...")
                request = service.files().get_media(fileId=file_id)
                with open(local_source_path, 'wb') as f:
                    downloader = MediaIoBaseDownload(f, request)
                    done = False
                    while not done: _, done = downloader.next_chunk()
                print("✅ Download complete.")
                file_ready = True
            else:
                print(f"❌ Error: File '{SOURCE_FILENAME}' not found.")
                list_folder_contents(service, FOLDER_ID)
                
        if not file_ready: return

        # 2. Process Raw Data
        print(f"🔄 Reading '{SOURCE_TAB_RAW}'...")
        xl = pd.ExcelFile(local_source_path)
        sheet_name = SOURCE_TAB_RAW if SOURCE_TAB_RAW in xl.sheet_names else "Raw"
        df = pd.read_excel(local_source_path, sheet_name=sheet_name, header=1)
        df = df.dropna(how='all', axis=1)

        if 'Entity' in df.columns:
            df['Entity'] = df['Entity'].astype(str).str.strip() 
        else:
            print("❌ Error: 'Entity' column not found!")
            return

        header_map = {normalize_header(col): col for col in df.columns}
        def find_col(k):
            t = normalize_header(k)
            for n, r in header_map.items(): 
                if t in n: return r
            return None

        col_exp = find_col("expensetoamortizeforperiod")
        col_proc = find_col("companyproceeds")
        col_vest = find_col("expensefairvalueofvestedawardsexercised") 
        col_early = find_col("expensefairvalueofearlyexercisedawardsvested")

        if not all([col_exp, col_proc, col_vest, col_early]):
            print("❌ Error: Missing required columns.")
            return

        sum_cols = [col_exp, col_proc, col_vest, col_early]

        df_filt = df[df['Entity'].isin(TARGET_ENTITIES)].copy()
        for col in sum_cols:
            df_filt[col] = pd.to_numeric(df_filt[col], errors='coerce').fillna(0)
        
        pivot_df = df_filt.groupby(GROUP_BY_COLS)[sum_cols].sum().reset_index()

        totals = pivot_df[sum_cols].sum()
        pivot_sums = {
            "Expense": float(totals.get(col_exp, 0.0)),
            "Proceeds": float(totals.get(col_proc, 0.0)),
            "Vested": float(totals.get(col_vest, 0.0)),
            "EarlyEx": float(totals.get(col_early, 0.0))
        }

        grand_row = {col: '' for col in pivot_df.columns}
        grand_row['Entity'] = 'Grand Total'
        for col in sum_cols: grand_row[col] = totals[col]
        pivot_df = pd.concat([pivot_df, pd.DataFrame([grand_row])], ignore_index=True)

        # 3. Process Journal Entries
        print(f"📖 Reading '{SOURCE_TAB_JE}'...")
        df_je = pd.read_excel(local_source_path, sheet_name=SOURCE_TAB_JE, header=1)
        je_map = {normalize_header(col): col for col in df_je.columns}
        def get_je_col(k):
            t = normalize_header(k)
            for n, r in je_map.items(): 
                if t in n: return r
            return None

        col_acc = get_je_col("expensetoamortize") or get_je_col("account") or df_je.columns[0]
        col_db = get_je_col("debit")
        col_cr = get_je_col("credit")

        df_calc = df_je.copy()
        df_calc[col_acc] = df_calc[col_acc].astype(str).str.strip()
        df_calc[col_db] = clean_currency(df_calc[col_db])
        df_calc[col_cr] = clean_currency(df_calc[col_cr])

        def sum_je(k, v_col):
            mask = df_calc[col_acc].str.contains(k, case=False, na=False)
            return df_calc.loc[mask, v_col].sum()

        sw_exp = sum_je("Compensation Expense", col_db)
        sw_proc = sum_je("Cash", col_db)
        mask_vest = df_calc[col_acc].str.contains(r"Paid[- ]in[- ]Capital", case=False, regex=True, na=False)
        sw_vest = df_calc.loc[mask_vest, col_db].sum()

        # 4. Prepare Entry Dataframes (Memory)
        def prepare_entry_df(entities, suffix):
            df_sub = pivot_df[
                (pivot_df['Entity'].isin(entities)) & 
                (pivot_df['Entity'] != 'Grand Total')
            ].copy()
            
            data = []
            total = df_sub[col_exp].sum()
            
            data.append({
                "Account": "30245 Equity : Additional Paid In Capital : APIC - Stock Option Compensation",
                "Debit": 0, "Credit": total,
                "Entity": entities[0] if entities else "", "Dept": "0000 Corporate",
                "Location": "1 San Francisco", "Class": "601 Horizontal",
                "Description": f"SBC Expense Accrual - {suffix}"
            })
            
            for _, row in df_sub.iterrows():
                amt = row[col_exp]
                if abs(amt) < 0.01: continue
                db, cr = (amt, 0) if amt > 0 else (0, abs(amt))
                data.append({
                    "Account": "60175 Personnel : Stock Option Compensation",
                    "Debit": db, "Credit": cr,
                    "Entity": row['Entity'], "Dept": row['Dept'],
                    "Location": row['Location'], "Class": row['Class'],
                    "Description": "SBC Expense Allocation"
                })
            
            df_out = pd.DataFrame(data)
            cols = ["Account", "Debit", "Credit", "Entity", "Dept", "Location", "Class", "Description"]
            for c in cols: 
                if c not in df_out.columns: df_out[c] = ""
            return df_out[cols]

        df_us_mem = prepare_entry_df(ENTRY_ENTITIES_US, "US")
        df_ca_mem = prepare_entry_df(ENTRY_ENTITIES_CA, "Canada")

        # 5. Write to Excel
        print(f"💾 Creating '{TARGET_FILENAME}'...")
        with pd.ExcelWriter(local_target_path, engine='openpyxl') as writer:
            pivot_df.to_excel(writer, sheet_name=PIVOT_SHEET_NAME, index=False)
            ws_piv = writer.sheets[PIVOT_SHEET_NAME]
            apply_header_style(ws_piv, 1, len(pivot_df.columns))
            auto_adjust_columns(ws_piv)

            pd.DataFrame().to_excel(writer, sheet_name=RECON_SHEET_NAME, index=False)
            ws_rec = writer.sheets[RECON_SHEET_NAME]
            
            headers = ["Category", "Per Pivot", "As per Shareworks", "Validation"]
            for i, h in enumerate(headers, 1):
                c = ws_rec.cell(row=2, column=i, value=h)
                c.font = BOLD_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER
                ws_rec.column_dimensions[get_column_letter(i)].width = 25
            
            r = 3
            def write_check(r, label, p, s):
                ws_rec.cell(r, 1, label).border = THIN_BORDER
                ws_rec.cell(r, 2, p).border = THIN_BORDER; ws_rec.cell(r, 2).number_format = '#,##0.00'
                ws_rec.cell(r, 3, s).border = THIN_BORDER; ws_rec.cell(r, 3).number_format = '#,##0.00'
                ws_rec.cell(r, 4, f'=IF(ABS(ROUND(B{r}-C{r}, 2))<1, "MATCH", "CHECK")').border = THIN_BORDER
                return r + 1
            r = write_check(r, "Expense", pivot_sums["Expense"], sw_exp)
            r = write_check(r, "Company Proceeds", pivot_sums["Proceeds"], sw_proc)
            r = write_check(r, "FV of Vested Option", pivot_sums["Vested"], sw_vest)
            # ==== Insert Missing INFO ONLY Row ====

            ws_rec.cell(r, 1, "Within Period (01-Oct-2025 to 31-Oct-2025)").border = THIN_BORDER
            ws_rec.cell(r, 1).font = BOLD_FONT
            r += 1  # move down
    
            ws_rec.cell(r, 1, "Expense (Fair Value) of Early Exercised Awards Vested in Period from Original Grant").border = THIN_BORDER
            ws_rec.cell(r, 2, pivot_sums["EarlyEx"]).border = THIN_BORDER
            ws_rec.cell(r, 2).number_format = '#,##0.00'
            ws_rec.cell(r, 3, "").border = THIN_BORDER
            ws_rec.cell(r, 4, "INFO ONLY").border = THIN_BORDER
            r += 2  # spacing before totals

            
            # ---- Restore: Write Shareworks / Expense (Journal Entries) Table ----
            jr = r + 3
            ws_rec.cell(jr - 1, 1, "Shareworks Entry (Expense Journal Entries)").font = BOLD_FONT

            # Write headers
            for i, col in enumerate(df_je.columns, 1):
                c = ws_rec.cell(jr, i, col)
                c.font = BOLD_FONT
                c.fill = HEADER_FILL
                c.border = THIN_BORDER

            # Write values
            for row_index, row in enumerate(df_je.values, 1):
                for col_index, value in enumerate(row, 1):
                    cell = ws_rec.cell(jr + row_index, col_index, value)
                    cell.border = THIN_BORDER

            # Add totals under Shareworks values
            total_row = jr + len(df_je) + 1
            ws_rec.cell(total_row, 1, "Grand Total").font = BOLD_FONT
            ws_rec.cell(total_row, 1).border = THIN_BORDER

            idx_debit = df_je.columns.get_loc(col_db) + 1
            idx_credit = df_je.columns.get_loc(col_cr) + 1

            ws_rec.cell(total_row, idx_debit, f"=SUM({get_column_letter(idx_debit)}{jr+1}:{get_column_letter(idx_debit)}{total_row-1})").border = THIN_BORDER
            ws_rec.cell(total_row, idx_credit, f"=SUM({get_column_letter(idx_credit)}{jr+1}:{get_column_letter(idx_credit)}{total_row-1})").border = THIN_BORDER

            
            df_us_mem.to_excel(writer, sheet_name=ENTRY_SHEET_NAME_US, index=False)
            ws_us = writer.sheets[ENTRY_SHEET_NAME_US]
            apply_header_style(ws_us, 1, len(df_us_mem.columns))
            auto_adjust_columns(ws_us)

            df_ca_mem.to_excel(writer, sheet_name=ENTRY_SHEET_NAME_CA, index=False)
            ws_ca = writer.sheets[ENTRY_SHEET_NAME_CA]
            apply_header_style(ws_ca, 1, len(df_ca_mem.columns))
            auto_adjust_columns(ws_ca)

        print("✅ Excel file created successfully.")
        
        # 6. Generate NetSuite Payloads (RELOAD LOGIC)
        print("⚙️ Generating NetSuite Payloads (Reading back from Excel)...")
        mappings = load_local_mappings(local_mapping_path)

        def balance_payload(payload):
            balanced_lines = []

            for line in payload["lines"]:
                balanced_lines.append(line)

                # Create mirrored entry identical except for swapped debit/credit
                mirror = line.copy()

                if line.get("debit", 0) > 0:
                    amount = line["debit"]
                    mirror["debit"] = 0
                    mirror["credit"] = amount
                elif line.get("credit", 0) > 0:
                    amount = line["credit"]
                    mirror["credit"] = 0
                    mirror["debit"] = amount
                else:
                    continue  # ignore zero lines

                mirror["memo"] = f"{line.get('memo', '')} (balance)"
                balanced_lines.append(mirror)

            payload["lines"] = balanced_lines
            return payload


        if mappings:
            try:
                print(f"   📂 Reading back '{ENTRY_SHEET_NAME_US}'...")
                df_us_final = pd.read_excel(local_target_path, sheet_name=ENTRY_SHEET_NAME_US).fillna(0)
                
                print(f"   📂 Reading back '{ENTRY_SHEET_NAME_CA}'...")
                df_ca_final = pd.read_excel(local_target_path, sheet_name=ENTRY_SHEET_NAME_CA).fillna(0)
                
                payload_us = generate_payload(df_us_final, 1, mappings, "US Payload")
                payload_ca = generate_payload(df_ca_final, 5, mappings, "Canada Payload")
                
                
                payload_us = balance_payload(payload_us)
                payload_ca = balance_payload(payload_ca)
                
                # Save JSONs locally FIRST (before updating DB)
                us_json_path = os.path.join(LOCAL_WORK_DIR, "SBC_Payload_US.json")
                ca_json_path = os.path.join(LOCAL_WORK_DIR, "SBC_Payload_CA.json")
                
                with open(us_json_path, 'w') as f:
                    json.dump(payload_us, f, indent=4)
                with open(ca_json_path, 'w') as f:
                    json.dump(payload_ca, f, indent=4)
                    
                print(f"✅ Payloads saved to '{us_json_path}' and '{ca_json_path}'.")

                # Update DB with PATHS
                update_agent_database(local_mapping_path, us_json_path, ca_json_path, local_target_path)

                # PRINT JSON OUTPUT TO CONSOLE
                print("\n" + "="*50)
                print("📝 GENERATED US PAYLOAD (Lines):")
                print("="*50)
                print(json.dumps(payload_us["lines"], indent=4))
                print("\n" + "="*50)
                print("📝 GENERATED CANADA PAYLOAD (Lines):")
                print("="*50)
                print(json.dumps(payload_ca["lines"], indent=4))
                print("="*50 + "\n")
                
            except Exception as e:
                print(f"❌ Error during payload generation/saving: {e}")
        else:
            print("⚠️ Skipping Payload generation: Mapping file not found.")

        print("🧹 Done.")

    except Exception as e:
        print(f"❌ Error: {e}")
    return {
        "period": period,
        "business_day": business_day,
        "tasks_received": len(tasks or []),
        "files_generated": [
            "Equity_JE_US.json"
        ],
        "approval_required": True
    }
if __name__ == "__main__":
    run_equity_agent(period="Manual Run")